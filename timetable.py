#!/usr/bin/python
#
#           _                                 _   _            
#          | |                               | | (_)           
#  _ __ ___| |_ _ __ ___  ___ _ __   ___  ___| |_ ___   _____  
# | '__/ _ \ __| '__/ _ \/ __| '_ \ / _ \/ __| __| \ \ / / _ \ 
# | | |  __/ |_| | | (_) \__ \ |_) |  __/ (__| |_| |\ V /  __/ 
# |_|  \___|\__|_|_ \___/|___/ .__/ \___|\___|\__|_| \_/ \___| 
# | |      | |   | |         | |                               
# | |_ __ _| |__ | | ___     |_|                               
# | __/ _` | '_ \| |/ _ \                                      
# | || (_| | |_) | |  __/                                      
#  \__\__,_|_.__/|_|\___|                                      
#                                                              
#                                                              

# script requirements
import sys
import argparse 
import os

# xslx requirements
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font

# insert date
from datetime import date

# CONSTANT : common path to spreadsheets : 
commonPath = "Licence/TimeTables/"


def main():
    args = parseArgs()
    color = parseColor()
    updateWorkbook(args, color)


def updateWorkbook(args, color):
    pathToWorkbook = commonPath+args.field+".xlsx"
    if os.path.exists(pathToWorkbook):
        workbook = readWorkbook(pathToWorkbook, args.subject)
    else:
        workbook = createWorkbook(args.subject)
    if existsSheet(workbook, args.subject):
        updateSheet(workbook[args.subject], args.course, color)
    else:
        updateSheet(workbook.create_sheet(args.subject), args.course, color)
    workbook.save(pathToWorkbook)


def updateSheet(sheet, course, color):
    for row in range(1, sheet.max_row+1):
        if sheet.cell(row, 1).value == course:
            addCell(sheet, row, color)
            return
    addCourse(sheet, course, color)


def addCourse(sheet, course, color):
    sheet.insert_rows(1)
    sheet.cell(1, 1).value = course
    addCell(sheet, 1, color)


def addCell(sheet, row, color):
    for column in range(2, sheet.max_column+1):
        if not(sheet.cell(row, column).value):
            addDate(sheet.cell(row,column), color)
            return
    addDate(sheet.cell(row, sheet.max_column+1), color)

def addDate(cell, color):
    today = date.today()
    formattedDate = today.strftime("%d/%m")
    cell.value = formattedDate
    cell.fill = PatternFill(fill_type="solid", fgColor=color) 
    cell.font = Font(bold = True)


def existsSheet(workbook, subject):
    return (subject in workbook.sheetnames)


def createWorkbook(subject):
    newWorkbook = Workbook()
    newWorksheet = newWorkbook.active
    newWorksheet.title = subject
    return newWorkbook


def readWorkbook(pathToWorkbook, subject):
    workbook = load_workbook(pathToWorkbook)
    return workbook


def parseArgs():
    parser = argparse.ArgumentParser(description='Update the subject retrospective time table.')
    parser.add_argument('field')
    parser.add_argument('subject')
    parser.add_argument('course')
    args = parser.parse_args()
    return args


def parseColor():
    print('ui')
    index = input("Confidence index ? <0 - 1 - 2> :")
    if(index == '2'): # Green
        return "99CC00"
    if(index == '1'): # Orange
        return "FFCC00"
    return "FF6600"


main()
